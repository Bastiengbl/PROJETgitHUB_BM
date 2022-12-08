
import openpyxl #openpyxl permet de lire, écrire et modifier des fichiers .xlsx
import random #Permet de generer des nombres aléatoire (notes élèves)
import os
import shutil

def Math():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
            
            
    #Calculer moyenne classe de la matière.
    sheet.cell(1, 5).value = 'Moyenne classe'
    MoyClass=cpt/nb_eleve
    sheet.cell(2, 5).value = (round(MoyClass,1))
    
        
        
    workbook.save('Math.xlsx')
    
    
def Reseau():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
    workbook.save('Reseau.xlsx')
     
def Anglais():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
    workbook.save('Anglais.xlsx')
    
    
def Francais():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
    workbook.save('Francais.xlsx')
    
def Physique():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
    workbook.save('Physique.xlsx')
    
def Telephonie():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
    workbook.save('Telephonie.xlsx')
    
def Web():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
    workbook.save('Web.xlsx')

def Linux():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
    workbook.save('Linux.xlsx')
    
def Programmation():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
    workbook.save('Programmation.xlsx')
    
print(Math(), Reseau(), Programmation(), Web(), Linux(), Telephonie(), Physique(), Francais(), Anglais()) #lance les def
os.chdir('/home/etudiant/PROJETgitHUB_BM/data/') #modifie l'emplacement afin de supprimer un potentiel fichier précédent
shutil.rmtree('Notes_matieres') #supprime un ancien fichier Notes(pou pouvoir relancer le progtramme)
os.mkdir('Notes_matieres') #création du fichier pour recenser les notes
shutil.move('/home/etudiant/PROJETgitHUB_BM/data/Anglais.xlsx', '/home/etudiant/PROJETgitHUB_BM/data/Notes_matieres/' )         #déplace les notes dans le dossier réservé.
shutil.move('/home/etudiant/PROJETgitHUB_BM/data/Francais.xlsx', '/home/etudiant/PROJETgitHUB_BM/data/Notes_matieres/' )
shutil.move('/home/etudiant/PROJETgitHUB_BM/data/Math.xlsx', '/home/etudiant/PROJETgitHUB_BM/data/Notes_matieres/' )
shutil.move('/home/etudiant/PROJETgitHUB_BM/data/Web.xlsx', '/home/etudiant/PROJETgitHUB_BM/data/Notes_matieres/' )
shutil.move('/home/etudiant/PROJETgitHUB_BM/data/Programmation.xlsx', '/home/etudiant/PROJETgitHUB_BM/data/Notes_matieres/' )
shutil.move('/home/etudiant/PROJETgitHUB_BM/data/Telephonie.xlsx', '/home/etudiant/PROJETgitHUB_BM/data/Notes_matieres/' )
shutil.move('/home/etudiant/PROJETgitHUB_BM/data/Physique.xlsx', '/home/etudiant/PROJETgitHUB_BM/data/Notes_matieres/' )
shutil.move('/home/etudiant/PROJETgitHUB_BM/data/Linux.xlsx', '/home/etudiant/PROJETgitHUB_BM/data/Notes_matieres/' )
shutil.move('/home/etudiant/PROJETgitHUB_BM/data/Reseau.xlsx', '/home/etudiant/PROJETgitHUB_BM/data/Notes_matieres/' )

    
    
    
    
    