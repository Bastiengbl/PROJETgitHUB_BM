import openpyxl #openpyxl permet de lire, Ã©crire et modifier des fichiers .xlsx
import random #Permet de generer des nombres alÃ©atoire (notes Ã©lÃ¨ves)
import os
import shutil

def notesS1():
    l=["Initiation_aux_reseaux_informatiques.xlsx",
"Principes_et_architecture_des_reseaux.xlsx",
"Reseaux_locaux_et_equipements_actifs.xlsx",
"Fondamentaux_des_Systemes_l_ectroniques.xlsx",
"Supports_de_transmission_pour_les_reseaux.xlsx",
"Architecture_des_systemes_numeriques_et_informatiques.xlsx",
"Fondamentaux_de_la_programmation.xlsx",
"Bases_des_systemes_d_exploitation.xlsx",
"Introduction_aux_technologies_Web.xlsx",
"Anglais_technique_1.xlsx",
"Expression_Culture_Communication_Professionnelles_Introduction_la_communication_et_au_savoir_etre_professionnels.xlsx",
"Projet_Personnel_et_Professionnel.xlsx",
"Mathematiques_du_signal.xlsx",
"Mathematiques_des_transmissions.xlsx",
"Gestion_de_projet_1_Maitriser_les_bases_de_l_organisation_du_travail.xlsx",
"Mathematiques_renforcement_1.xlsx",
"Se_Senbiliser_a_l_hygiene_informatique_et_a_la_cybersecurite.xlsx",
"S_initier_aux_reseaux_informatiques.xlsx",
"Decouvrir_un_dispositif_de_transmission.xlsx",
"Se_presenter_sur_Internet.xlsx",
"Traiter_des_donnees.xlsx",
"Demarche_Portfolio.xlsx"
]
    for y in l:
        
        workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des Ã©lÃ¨ves a tout moment.
        sheet_ranges = workbook['Feuille'] 
    
        #Variables
        nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupÃ©rer le nombre total d'Ã©lÃ¨ves.
        sheet = workbook.active
    
    
    
        #Mettre notes alÃ©atoires dans la colonne note.
        sheet.cell(1, 4).value = 'Note'
        for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
            i=2
            cpt=0 #Permet de faire la moyenne des notes par la suite.
            while i<=nb_eleve:
                rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de gÃ©nÃ©rer un nombre alÃ©atoire.
                sheet.cell(i, 4).value = rdm_note
                cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
                i=i+1
                 #Calculer moyenne classe de la matiÃ¨re.
        sheet.cell(1, 5).value = 'Moyenne classe'
        MoyClass=cpt/nb_eleve
        sheet.cell(2, 5).value = (round(MoyClass,1))
        workbook.save(y)

notesS1()


def notesS2():
    l=["Technologie_de_l_Internet.xlsx",
"Administration_systeme_et_fondamentaux_de_la_virtualisation.xlsx",
"Bases_des_services_reseaux.xlsx",
"Initiation_a_la_telephonie_d_entreprise.xlsx",
"Signaux_et_Systemes_pour_les_transmissions.xlsx",
"Numerisation_de_l_information.xlsx",
"Sources_de_donnees.xlsx",
"Analyse_et_traitement_de_donnees_structurees.xlsx",
"Initiation_au_developpement_Web.xlsx",
"Anglais_technique_2.xlsx",
"Expression_Culture_Communication_Professionnelles_Renforcement_des_techniques_de_communication.xlsx",
"Projet_Personnel_et_Professionnel.xlsx",
"Mathematiques_des_systemes_numeriques.xlsx",
"Analyse_mathematique_des_signaux.xlsx",
"Mathematiques_renforcement_2.xlsx",
"Construire_un_reseau_informatique_pour_une_petite_structure.xlsx",
"Mesurer_et_caracteriser_un_signal_ou_un_systeme.xlsx",
"Mettre_en_place_une_solution_informatique_pour_l_entreprise.xlsx",
"Projet_integratif.xlsx",
"Demarche_Portfolio.xlsx"]
    
    for y in l:
        
        workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des Ã©lÃ¨ves a tout moment.
        sheet_ranges = workbook['Feuille'] 
    
        #Variables
        nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupÃ©rer le nombre total d'Ã©lÃ¨ves.
        sheet = workbook.active
    
    
    
        #Mettre notes alÃ©atoires dans la colonne note.
        sheet.cell(1, 4).value = 'Note'
        for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
            i=2
            cpt=0 #Permet de faire la moyenne des notes par la suite.
            while i<=nb_eleve:
                rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de gÃ©nÃ©rer un nombre alÃ©atoire.
                sheet.cell(i, 4).value = rdm_note
                cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
                i=i+1
                 #Calculer moyenne classe de la matiÃ¨re.
        sheet.cell(1, 5).value = 'Moyenne classe'
        MoyClass=cpt/nb_eleve
        sheet.cell(2, 5).value = (round(MoyClass,1))
        workbook.save(y)

notesS2()




