import openpyxl
 
coefT=100
 
 
 
def UE1_S1(Nom_):
    Nom_=Nom_.upper()
    Moyenne1=0


    liste_matière_coef2=["notes_S1/Projet_Personnel_et_Professionnel.xlsx","notes_S1/Mathematiques_renforcement_1.xlsx"]
    liste_matière_coef4=["notes_S1/Architecture_des_systemes_numeriques_et_informatiques.xlsx","notes_S1/Anglais_technique_1.xlsx","notes_S1/Expression_Culture_Communication_Professionnelles_Introduction_la_communication_et_au_savoir_etre_professionnels.xlsx","notes_S1/Mathematiques_du_signal.xlsx","notes_S1/Mathematiques_des_transmissions.xlsx"]
    liste_matière_coef5=["notes_S1/Bases_des_systemes_d_exploitation.xlsx"]
    liste_matière_coef6=["notes_S1/Reseaux_locaux_et_equipements_actifs.xlsx","notes_S1/Fondamentaux_des_Systemes_l_ectroniques.xlsx"]
    liste_matière_coef9=["notes_S1/Initiation_aux_reseaux_informatiques.xlsx","notes_S1/Principes_et_architecture_des_reseaux.xlsx"]
    liste_matière_coef11=["notes_S1/Se_Senbiliser_a_l_hygiene_informatique_et_a_la_cybersecurite.xlsx"]
    liste_matière_coef30=["notes_S1/S_initier_aux_reseaux_informatiques.xlsx"]

    for fichier in liste_matière_coef2:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*2
                Moyenne1=Moyenne1+cellule_adjacente_
                      
    for fichier in liste_matière_coef4:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*4
                Moyenne1=Moyenne1+cellule_adjacente_

    for fichier in liste_matière_coef5:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*5
                Moyenne1=Moyenne1+cellule_adjacente_
 
    for fichier in liste_matière_coef6:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*6
                Moyenne1=Moyenne1+cellule_adjacente_
 
    for fichier in liste_matière_coef9:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*9
                Moyenne1=Moyenne1+cellule_adjacente_
 
    for fichier in liste_matière_coef11:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*11
                Moyenne1=Moyenne1+cellule_adjacente_
                             
    for fichier in liste_matière_coef30:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*30
                Moyenne1=Moyenne1+cellule_adjacente_
                
    Moyenne1T=Moyenne1/coefT
    Moyenne1T=round(Moyenne1T,2)
    return(Moyenne1T)

def UE2_S1(Nom_):
    Nom_=Nom_.upper()
    Moyenne2=0

    liste_matière_coef2=["notes_S1/Projet_Personnel_et_Professionnel.xlsx","notes_S1/Mathematiques_renforcement_1.xlsx","notes_S1/Reseaux_locaux_et_equipements_actifs.xlsx"]
    liste_matière_coef3=["notes_S1/Gestion_de_projet_1_Maitriser_les_bases_de_l_organisation_du_travail.xlsx"]
    liste_matière_coef4=["notes_S1/Initiation_aux_reseaux_informatiques.xlsx"]
    liste_matière_coef6=["notes_S1/Expression_Culture_Communication_Professionnelles_Introduction_la_communication_et_au_savoir_etre_professionnels.xlsx","notes_S1/Anglais_technique_1.xlsx"]
    liste_matière_coef7=["notes_S1/Supports_de_transmission_pour_les_reseaux.xlsx"]
    liste_matière_coef9=["notes_S1/Mathematiques_du_signal.xlsx","notes_S1/Mathematiques_des_transmissions.xlsx"]
    liste_matière_coef10=["notes_S1/Fondamentaux_des_Systemes_l_ectroniques.xlsx"]
    liste_matière_coef40=["notes_S1/Decouvrir_un_dispositif_de_transmission.xlsx"]
    
    for fichier in liste_matière_coef2:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*2
                Moyenne2=Moyenne2+cellule_adjacente_
                     
    for fichier in liste_matière_coef3:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*3
                Moyenne2=Moyenne2+cellule_adjacente_
 
    for fichier in liste_matière_coef4:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*4
                Moyenne2=Moyenne2+cellule_adjacente_
 
    for fichier in liste_matière_coef6:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*6
                Moyenne2=Moyenne2+cellule_adjacente_
 
    for fichier in liste_matière_coef7:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*7
                Moyenne2=Moyenne2+cellule_adjacente_
 
    for fichier in liste_matière_coef9:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*9
                Moyenne2=Moyenne2+cellule_adjacente_
                
    for fichier in liste_matière_coef10:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*10
                Moyenne2=Moyenne2+cellule_adjacente_
                             
    for fichier in liste_matière_coef40:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*40
                Moyenne2=Moyenne2+cellule_adjacente_
                         
    Moyenne2T=Moyenne2/coefT
    Moyenne2T=round(Moyenne2T,2)
    return(Moyenne2T)

def UE3_S1(Nom_):
    Nom_=Nom_.upper()
    Moyenne3=0
    
    liste_matière_coef2=["notes_S1/Projet_Personnel_et_Professionnel.xlsx","notes_S1/Reseaux_locaux_et_equipements_actifs.xlsx"]
    liste_matière_coef3=["notes_S1/Gestion_de_projet_1_Maitriser_les_bases_de_l_organisation_du_travail.xlsx"]
    liste_matière_coef4=["notes_S1/Initiation_aux_reseaux_informatiques.xlsx","notes_S1/Gestion_de_projet_1_Maitriser_les_bases_de_l_organisation_du_travail.xlsx"]
    liste_matière_coef5=["notes_S1/Expression_Culture_Communication_Professionnelles_Introduction_la_communication_et_au_savoir_etre_professionnels.xlsx","notes_S1/Architecture_des_systemes_numeriques_et_informatiques.xlsx","notes_S1/Anglais_technique_1.xlsx"]
    liste_matière_coef6=["notes_S1/Bases_des_systemes_d_exploitation.xlsx"]
    liste_matière_coef9=["notes_S1/Se_presenter_sur_Internet.xlsx"]
    liste_matière_coef19=["notes_S1/Fondamentaux_de_la_programmation.xlsx"]
    liste_matière_coef36=["notes_S1/Traiter_des_donnees.xlsx"]
        
    for fichier in liste_matière_coef2:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*2
                Moyenne3=Moyenne3+cellule_adjacente_
                   
    for fichier in liste_matière_coef3:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*3
                Moyenne3=Moyenne3+cellule_adjacente_

    for fichier in liste_matière_coef4:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*4
                Moyenne3=Moyenne3+cellule_adjacente_

    for fichier in liste_matière_coef5:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*5
                Moyenne3=Moyenne3+cellule_adjacente_
 
    for fichier in liste_matière_coef6:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*6
                Moyenne3=Moyenne3+cellule_adjacente_
 
    for fichier in liste_matière_coef9:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*9
                Moyenne3=Moyenne3+cellule_adjacente_
                    
    for fichier in liste_matière_coef19:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*19
                Moyenne3=Moyenne3+cellule_adjacente_
 
    for fichier in liste_matière_coef36:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*36
                Moyenne3=Moyenne3+cellule_adjacente_
                         
    Moyenne3T=Moyenne3/coefT
    Moyenne3T=round(Moyenne3T,2)
    return(Moyenne3T)
 

def UEt_S1(Nom_):
    Moyenne1_S1=UE1_S1(Nom_)
    Moyenne2_S1=UE2_S1(Nom_)
    Moyenne3_S1=UE3_S1(Nom_)
    MoyenneT_S1=((Moyenne1_S1+Moyenne2_S1+Moyenne3_S1)/3)
    MoyenneT_S1=(round(MoyenneT_S1,2))
    return MoyenneT_S1






 
def UE1_S2(Nom_):
    Nom_=Nom_.upper()
    Moyenne1=0

    liste_matière_coef1=["notes_S2/Sources_de_donnees.xlsx"]
    liste_matière_coef2=["notes_S2/Demarche_Portfolio.xlsx","notes_S2/Mathematiques_renforcement_2.xlsx","notes_S2/Analyse_mathematique_des_signaux.xlsx","notes_S2/Projet_Personnel_et_Professionnel.xlsx","notes_S2/Initiation_au_developpement_Web.xlsx"]
    liste_matière_coef3=["notes_S2/Initiation_a_la_telephonie_d_entreprise.xlsx","notes_S2/Mathematiques_des_systemes_numeriques.xlsx","notes_S2/Numerisation_de_l_information.xlsx"]
    liste_matière_coef4=["notes_S2/Signaux_et_Systemes_pour_les_transmissions.xlsx","notes_S2/Expression_Culture_Communication_Professionnelles_Renforcement_des_techniques_de_communication.xlsx"]
    liste_matière_coef6=["notes_S2/Administration_systeme_et_fondamentaux_de_la_virtualisation.xlsx"]
    liste_matière_coef7=["notes_S2/Anglais_technique_2.xlsx"]
    liste_matière_coef9=["notes_S2/Technologie_de_l_Internet.xlsx","notes_S2/Bases_des_services_reseaux.xlsx"]
    liste_matière_coef17=["notes_S2/Projet_integratif.xlsx"]
    liste_matière_coef24=["notes_S2/Construire_un_reseau_informatique_pour_une_petite_structure.xlsx"]
    
        
    for fichier in liste_matière_coef1:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*1
                Moyenne1=Moyenne1+cellule_adjacente_
                 
    for fichier in liste_matière_coef2:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*2
                Moyenne1=Moyenne1+cellule_adjacente_
                 
    for fichier in liste_matière_coef3:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*3
                Moyenne1=Moyenne1+cellule_adjacente_

    for fichier in liste_matière_coef4:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*4
                Moyenne1=Moyenne1+cellule_adjacente_
 
    for fichier in liste_matière_coef6:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*6
                Moyenne1=Moyenne1+cellule_adjacente_
 
    for fichier in liste_matière_coef7:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*7
                Moyenne1=Moyenne1+cellule_adjacente_
 
    for fichier in liste_matière_coef9:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*9
                Moyenne1=Moyenne1+cellule_adjacente_
                
    for fichier in liste_matière_coef17:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*17
                Moyenne1=Moyenne1+cellule_adjacente_
                             
    for fichier in liste_matière_coef24:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*24
                Moyenne1=Moyenne1+cellule_adjacente_
                         
    Moyenne1T=Moyenne1/coefT
    Moyenne1T=round(Moyenne1T,2)
    return(Moyenne1T)

def UE2_S2(Nom_):
    Nom_=Nom_.upper()
    Moyenne2=0
    
    liste_matière_coef2=["notes_S2/Projet_Personnel_et_Professionnel.xlsx","notes_S2/Demarche_Portfolio.xlsx","notes_S2/Mathematiques_renforcement_2.xlsx"]
    liste_matière_coef3=["notes_S2/Mathematiques_des_systemes_numeriques.xlsx"]
    liste_matière_coef5=["notes_S2/Expression_Culture_Communication_Professionnelles_Renforcement_des_techniques_de_communication.xlsx"]
    liste_matière_coef6=["notes_S2/Technologie_de_l_Internet.xlsx","notes_S2/Initiation_a_la_telephonie_d_entreprise.xlsx"]
    liste_matière_coef7=["notes_S2/Anglais_technique_2.xlsx"]
    liste_matière_coef8=["notes_S2/Numerisation_de_l_information.xlsx","notes_S2/Analyse_mathematique_des_signaux.xlsx"]
    liste_matière_coef11=["notes_S2/Signaux_et_Systemes_pour_les_transmissions.xlsx"]
    liste_matière_coef17=["notes_S2/Projet_integratif.xlsx"]
    liste_matière_coef23=["notes_S2/Mesurer_et_caracteriser_un_signal_ou_un_systeme.xlsx"]
        
    for fichier in liste_matière_coef2:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*2
                Moyenne2=Moyenne2+cellule_adjacente_
                 
    for fichier in liste_matière_coef3:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*3
                Moyenne2=Moyenne2+cellule_adjacente_
                         
    for fichier in liste_matière_coef5:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*4
                Moyenne2=Moyenne2+cellule_adjacente_
 
    for fichier in liste_matière_coef6:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_  
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*6
                Moyenne2=Moyenne2+cellule_adjacente_
 
    for fichier in liste_matière_coef7:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*7
                Moyenne2=Moyenne2+cellule_adjacente_
 
    for fichier in liste_matière_coef8:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*8
                Moyenne2=Moyenne2+cellule_adjacente_
 
    for fichier in liste_matière_coef11:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*11
                Moyenne2=Moyenne2+cellule_adjacente_
                     
    for fichier in liste_matière_coef17:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*17
                Moyenne2=Moyenne2+cellule_adjacente_
                             
    for fichier in liste_matière_coef23:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*23
                Moyenne2=Moyenne2+cellule_adjacente_
                                   
    Moyenne2T=Moyenne2/coefT
    Moyenne2T=round(Moyenne2T,2)
    return(Moyenne2T)

def UE3_S2(Nom_):
    Nom_=Nom_.upper()
    Moyenne3=0
    
    liste_matière_coef2=["notes_S2/Projet_Personnel_et_Professionnel.xlsx","notes_S2/Demarche_Portfolio.xlsx"]
    liste_matière_coef3=["notes_S2/Mathematiques_renforcement_2.xlsx"]
    liste_matière_coef4=["notes_S2/Anglais_technique_2.xlsx","notes_S2/Expression_Culture_Communication_Professionnelles_Renforcement_des_techniques_de_communication.xlsx"]
    liste_matière_coef5=["notes_S2/Technologie_de_l_Internet.xlsx","notes_S2/Initiation_a_la_telephonie_d_entreprise.xlsx"]
    liste_matière_coef6=["notes_S2/Administration_systeme_et_fondamentaux_de_la_virtualisation.xlsx","notes_S2/Mathematiques_des_systemes_numeriques.xlsx"]     
    liste_matière_coef7=["notes_S2/Sources_de_donnees.xlsx","notes_S2/Analyse_et_traitement_de_donnees_structurees.xlsx"]
    liste_matière_coef9=["notes_S2/Initiation_au_developpement_Web.xlsx"]
    liste_matière_coef16=["notes_S2/Projet_integratif.xlsx"]
    liste_matière_coef24=["notes_S2/Mettre_en_place_une_solution_informatique_pour_l_entreprise.xlsx"]

    for fichier in liste_matière_coef2:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*2
                Moyenne3=Moyenne3+cellule_adjacente_
                 
    for fichier in liste_matière_coef3:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*3
                Moyenne3=Moyenne3+cellule_adjacente_

    for fichier in liste_matière_coef4:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*4
                Moyenne3=Moyenne3+cellule_adjacente_
 
    for fichier in liste_matière_coef5:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*5
                Moyenne3=Moyenne3+cellule_adjacente_
 
    for fichier in liste_matière_coef6:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*6
                Moyenne3=Moyenne3+cellule_adjacente_
 
    for fichier in liste_matière_coef7:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*7
                Moyenne3=Moyenne3+cellule_adjacente_
                
    for fichier in liste_matière_coef9:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*9
                Moyenne3=Moyenne3+cellule_adjacente_
 
    for fichier in liste_matière_coef16:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*16
                Moyenne3=Moyenne3+cellule_adjacente_
                
    for fichier in liste_matière_coef24:
        wb = openpyxl.load_workbook(fichier)
        sheet = wb['Feuille']
        texte_a_rechercher = Nom_
        for row in sheet.rows:
            cell = row[2]
            if cell.value == texte_a_rechercher:
                cellule_adjacente = sheet.cell(row=cell.row, column=cell.col_idx + 1)
                cellule_adjacente_ = float(cellule_adjacente.value)
                cellule_adjacente_=cellule_adjacente_*24
                Moyenne3=Moyenne3+cellule_adjacente_
                                    
    Moyenne3T=Moyenne3/coefT
    Moyenne3T=round(Moyenne3T,2)
    return(Moyenne3T)
             

def UEt_S2(Nom_):
    Moyenne1_S2=UE1_S2(Nom_)
    Moyenne2_S2=UE2_S2(Nom_)
    Moyenne3_S2=UE3_S2(Nom_)
    MoyenneT_S2=((Moyenne1_S2+Moyenne2_S2+Moyenne3_S2)/3)
    MoyenneT_S2=(round(MoyenneT_S2,2))
    return MoyenneT_S2
    Moyenne1_S2=UE1_S2(Nom_)
    Moyenne2_S2=UE2_S2(Nom_)
    Moyenne3_S2=UE3_S2(Nom_)
    MoyenneT_S2=((Moyenne1_S2+Moyenne2_S2+Moyenne3_S2)/3)
    MoyenneT_S2=(round(MoyenneT_S2,2))
    return MoyenneT_S2










def UE(Nom_):
    
    Nom_=Nom_.upper()
    Moyenne_S1=UEt_S1(Nom_)
    Moyenne_S2=UEt_S2(Nom_)
    Moyenne=(Moyenne_S1+Moyenne_S2)/2
    Moyenne=round(Moyenne,2)
    return(Moyenne)
    





        
        
def UE_Moy_T():
    Nom_=""
    som_moyenne=0
    moyenne_totale=0
    wb = openpyxl.load_workbook("Nom_Prenom.xlsx")
    sheet = wb['Feuille']
    for row in sheet.rows:
        cell = row[2]
        Nom_=cell.value
        moyenne=UE(Nom_)
        moyenne=int(moyenne or 0) # Convertie en int
        som_moyenne +=moyenne
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des Ã©lÃ¨ves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    #Variables
    nb_eleve = (sheet_ranges['A2'].value) #Permet de recupÃ©rer le nombre total d'Ã©lÃ¨ves.
    sheet = workbook.active
    moyenne_totale += som_moyenne/nb_eleve
    moyenne_totale=round(moyenne_totale,2)
    print(moyenne_totale)
    
    
    
def ultime():
    