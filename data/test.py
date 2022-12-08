import openpyxl #openpyxl permet de lire, écrire et modifier des fichiers .xlsx
import random #Permet de generer des nombres aléatoire (notes élèves)




#Fonction page excel par matière.
def Math():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
            
            
    #Calculer moyenne classe de la matière.
    sheet.cell(1, 5).value = 'Moyenne classe'
    MoyClass=cpt/nb_eleve
    sheet.cell(2, 5).value = (round(MoyClass,1))
    
        
        
    workbook.save('Mathematiques.xlsx')
    
    
def Reseau():
    workbook = openpyxl.load_workbook('Nom_Prenom.xlsx') #Ouvre le fichier "Nom_Prenom.xlsx", il est donc possible de modifier les noms des élèves a tout moment.
    sheet_ranges = workbook['Feuille'] 
    
    #Variables
    nb_eleve = (sheet_ranges['A2'].value)+1 #Permet de recupérer le nombre total d'élèves.
    sheet = workbook.active
    
    
    
    #Mettre notes aléatoires dans la colonne note.
    sheet.cell(1, 4).value = 'Note'
    for row in sheet.iter_cols(min_col = 4, max_col = 4): #Boucles afin de remplir la colonne note. 
        i=2
        cpt=0 #Permet de faire la moyenne des notes par la suite.
        while i<=nb_eleve:
            rdm_note = round(random.uniform(0.0, 20.0), 1) #Permet de générer un nombre aléatoire.
            sheet.cell(i, 4).value = rdm_note
            cpt=cpt+rdm_note #Permet de faire la moyenne des notes par la suite.
            i=i+1
            
            
    #Calculer moyenne classe de la matière.
    sheet.cell(1, 5).value = 'Moyenne classe'
    MoyClass=(cpt/nb_eleve)
    sheet.cell(2, 5).value = (round(MoyClass,1))
    
        
        
    workbook.save('reseau.xlsx')
    
    
    
   
    
print(Math(),Reseau())
